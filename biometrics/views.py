# views.py in biometrics app

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import json
import csv
import os
from django.conf import settings
from django.shortcuts import render
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db
from django.http import HttpResponse
import io
import json
import openpyxl
from openpyxl.utils import get_column_letter


# Optional: Create a directory to store CSVs if not exists
CSV_DIR = os.path.join(settings.BASE_DIR, 'biometric_data')
os.makedirs(CSV_DIR, exist_ok=True)



cred = credentials.Certificate('/etc/secrets/firebase_key.json')


firebase_admin.initialize_app(cred, {
    'databaseURL': 'https://behavioural-auth-default-rtdb.firebaseio.com/'
})


def index(request):
    return render(request, 'index.html')

def export_page(request):
    return render(request, 'export.html')

@csrf_exempt
def save_behavior(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            user = data.get('user', 'user1')
            gender = data.get('gender', 'M')

            filename = f"{user}.csv"
            filepath = os.path.join(CSV_DIR, filename)

            # Check if file exists to write header
            file_exists = os.path.isfile(filepath)

            with open(filepath, 'a', newline='') as csvfile:
                fieldnames = [
                    'cpm', 'error_rate', 'dwell_avg', 'dwell_std',
                    'flight_avg', 'flight_std', 'click_dwell_avg',
                    'click_flight_avg', 'pressure_touch', 'scroll_x',
                    'scroll_y', 'double_click', 'swipe_speed_avg',
                    'tilt_angle_avg', 'gender'
                ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                if not file_exists:
                    writer.writeheader()

                data['gender'] = gender
                writer.writerow({field: data.get(field, '') for field in fieldnames})
            store_behavioral_data_to_firebase(user, gender, data)    

            return JsonResponse({'status': 'success', 'message': 'Data saved to CSV and Firebase successfully.'})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    return JsonResponse({'status': 'error', 'message': 'Invalid request method.'}, status=405)

def store_behavioral_data_to_firebase(user_id, gender, data_dict):
    """
    Push data to Firebase under each user.
    """
    ref = db.reference(f"/users/{user_id}")
    ref.push({
        "gender": gender,
        "data": data_dict
    })
    
def firebase_summary(request):
    ref = db.reference("users")
    users_data = ref.get()

    summary = []
    for user_id, records in (users_data or {}).items():
        summary.append({
            "user": user_id,
            "count": len(records)
        })

    return render(request, 'summary.html', {"summary": summary})

def export_csv(request):
    # Reference to all users in Firebase
    ref = db.reference("/users")
    users_data = ref.get()

    if not users_data:
        return HttpResponse("No data to export.", content_type="text/plain")

    # Define response as CSV
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="firebase_export.csv"'

    writer = csv.writer(response)

    # Write headers
    headers = [
        'user', 'gender', 'cpm', 'error_rate', 'dwell_avg', 'dwell_std',
        'flight_avg', 'flight_std', 'click_dwell_avg', 'click_flight_avg',
        'pressure_touch', 'scroll_x', 'scroll_y', 'double_click',
        'swipe_speed_avg', 'tilt_angle_avg'
    ]
    writer.writerow(headers)

    for user_id, entries in users_data.items():
        for entry_id, entry in entries.items():
            gender = entry.get('gender', '')
            data = entry.get('data', {})
            row = [
                user_id,
                gender,
                data.get('cpm', ''),
                data.get('error_rate', ''),
                data.get('dwell_avg', ''),
                data.get('dwell_std', ''),
                data.get('flight_avg', ''),
                data.get('flight_std', ''),
                data.get('click_dwell_avg', ''),
                data.get('click_flight_avg', ''),
                data.get('pressure_touch', ''),
                data.get('scroll_x', ''),
                data.get('scroll_y', ''),
                data.get('double_click', ''),
                data.get('swipe_speed_avg', ''),
                data.get('tilt_angle_avg', '')
            ]
            writer.writerow(row)

    return response

def export_json(request):
    ref = db.reference('/users/')
    data = ref.get()

    response = JsonResponse(data, safe=False)
    response['Content-Disposition'] = 'attachment; filename="export.json"'
    return response

def export_xlsx(request):
    ref = db.reference('/users/')
    data = ref.get()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for user, entries in data.items():
        ws = wb.create_sheet(title=user)
        ws.append(['Gender', 'Data Key', 'Value'])

        for entry_id, entry_data in entries.items():
            gender = entry_data.get('gender', 'N/A')
            for k, v in entry_data['data'].items():
                ws.append([gender, k, v])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
    wb.save(response)
    return response